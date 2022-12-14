import openpyxl

# Exercise one Find all companies and how many products they produce
# Loads workbook and all files
inventory_file = openpyxl.load_workbook('inventory.xlsx')
product_list = inventory_file['Sheet1']

# Calculate amount of products per supplier, set Dictionary to empty
products_per_supplier = {}
total_value_per_supplier = {}
products_inventory_under_ten = {}
# range allows for rows ot be looped through starting at 2nd row to skio descriptor row

# for every row in product_list from row 2 to last row
for product_row in range(2, product_list.max_row + 1):
    # grab name from product_row column 4
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    # if name already exists set value to current_num_products
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        # then increase value by 1
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        # Adds new supplier to dicitonary and sets value to one
        products_per_supplier[supplier_name] = 1

    # Calculate total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Calculate products inventory less than 10 product_num is key / inventory is value
    if inventory < 10:
        products_inventory_under_ten[product_num] = inventory

    # add value for total inventory price
    inventory_price.value = inventory * price


print(products_per_supplier)
print(total_value_per_supplier)
print(products_inventory_under_ten)

# creates a new file with updated values after program finishes running
inventory_file.save('inventory_with_total_value.xlsx')

